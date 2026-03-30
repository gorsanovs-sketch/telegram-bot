from pathlib import Path
from openpyxl import load_workbook

from telegram import ReplyKeyboardMarkup, Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

BOT_TOKEN = "8621874021:AAF7nHGqnL7LGI9sQlxSZATrany7SX5pRtE"
AUTOPARK_PIN = "4990"

ROOT_DIR = Path(r"C:\Users\Andris G\Desktop\Telegramm bot")
BASE_DIR = Path(__file__).parent

INSTR_DIR = BASE_DIR / "instrukcijas"
FORMS_DIR = ROOT_DIR / "Veidlapas"
AUTOPARK_DIR = ROOT_DIR / "Tehniķu autoparks"

MAIN_MENU = ReplyKeyboardMarkup(
    [
        ["Apsardzes centrāles instrukcija", "Tehniķa check-list"],
        ["Pults operatora darbības", "Kontakti", "Veidlapas"],
        ["Autoparks"],
    ],
    resize_keyboard=True,
    is_persistent=True,
)

BACK_MENU = ReplyKeyboardMarkup(
    [["⬅️ Atpakaļ uz galveno izvēlni"]],
    resize_keyboard=True,
    is_persistent=True,
)

CONTACTS_MENU = ReplyKeyboardMarkup(
    [
        ["🟢 Tehniskais dienests", "🔵 Pults EVOR AB"],
        ["🟡 Pults Evor Apsardze", "🟣 Klientu daļa"],
        ["🔴 Pults EvorM"],
        ["⬅️ Atpakaļ uz galveno izvēlni"],
    ],
    resize_keyboard=True,
    is_persistent=True,
)

CONTACTS = {
    "🟢 Tehniskais dienests": "+37129520000",
    "🔵 Pults EVOR AB": "+37120825555",
    "🟡 Pults Evor Apsardze": "+37127792000",
    "🟣 Klientu daļa": "+37129250000",
    "🔴 Pults EvorM": "+37122049665",
}


def get_files(folder: Path):
    if not folder.exists():
        return []

    allowed = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}
    return sorted(
        [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in allowed],
        key=lambda x: x.name.lower()
    )


def build_menu(files):
    rows = []
    row = []

    for f in files:
        row.append(f.stem[:40])
        if len(row) == 2:
            rows.append(row)
            row = []

    if row:
        rows.append(row)

    rows.append(["⬅️ Atpakaļ uz galveno izvēlni"])

    return ReplyKeyboardMarkup(
        rows,
        resize_keyboard=True,
        is_persistent=True,
    )


def build_map(files):
    return {f.stem[:40]: f for f in files}


def get_autopark_file():
    if not AUTOPARK_DIR.exists():
        return None

    files = sorted(
        [f for f in AUTOPARK_DIR.iterdir() if f.is_file() and f.suffix.lower() in {".xlsx", ".xls"}],
        key=lambda x: x.name.lower()
    )
    return files[0] if files else None


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def get_autopark_data():
    autopark_file = get_autopark_file()

    if not autopark_file:
        return None, None, f"❌ Nav Excel faila mapē:\n{AUTOPARK_DIR}"

    try:
        wb = load_workbook(autopark_file, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return None, None, "❌ Excel failā nav datu."

        data_rows = rows[1:]

        tech_info_map = {}
        button_map = {}

        for row in data_rows:
            if not row or len(row) < 2:
                continue

            # B aile = tehniķa vārds/uzvārds
            tech_name = clean_value(row[1])
            if not tech_name:
                continue

            info_lines = []

            # Rāda tikai no C ailes uz priekšu
            for value in row[2:]:
                value_text = clean_value(value)
                if value_text:
                    info_lines.append(value_text)

            if not info_lines:
                info_lines.append("Nav papildu informācijas.")

            tech_info_map[tech_name] = "\n".join(info_lines)
            button_map[tech_name[:40]] = tech_name

        if not tech_info_map:
            return None, None, "❌ Neizdevās atrast tehniķus B ailē."

        return tech_info_map, button_map, None

    except Exception as e:
        return None, None, f"❌ Kļūda lasot Excel:\n{e}"


def build_autopark_menu(button_map):
    rows = []
    row = []

    for btn in button_map.keys():
        row.append(btn)
        if len(row) == 2:
            rows.append(row)
            row = []

    if row:
        rows.append(row)

    rows.append(["⬅️ Atpakaļ uz galveno izvēlni"])

    return ReplyKeyboardMarkup(
        rows,
        resize_keyboard=True,
        is_persistent=True,
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Sveicināti! Izvēlieties sadaļu:",
        reply_markup=MAIN_MENU,
    )


async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    chat_id = update.effective_chat.id

    instr_files = get_files(INSTR_DIR)
    instr_map = build_map(instr_files)

    form_files = get_files(FORMS_DIR)
    form_map = build_map(form_files)

    tech_info_map, button_map, autopark_error = get_autopark_data()

    # Inicializējam piekļuves stāvokli
    if "autopark_access" not in context.chat_data:
        context.chat_data["autopark_access"] = False

    if "waiting_for_autopark_pin" not in context.chat_data:
        context.chat_data["waiting_for_autopark_pin"] = False

    if text == "Apsardzes centrāles instrukcija":
        if not instr_files:
            await update.message.reply_text(
                "Nav instrukciju.",
                reply_markup=MAIN_MENU,
            )
            return

        await update.message.reply_text(
            "Izvēlieties instrukciju:",
            reply_markup=build_menu(instr_files),
        )
        return

    if text in instr_map:
        with open(instr_map[text], "rb") as f:
            await update.message.reply_document(
                document=f,
                caption=f"Instrukcija: {text}",
                reply_markup=BACK_MENU,
            )
        return

    if text == "Tehniķa check-list":
        await update.message.reply_text(
            "TEHNIĶA CHECK-LIST\n\n"
            "1. Uztaisīt foto ar uzlīmēm\n"
            "2. Pārbaudīt vai ir foto ar fasādi,keypad,paneli\n"
            "3. Pārbaudīt trauksmes signālus\n"
            "4. Ierakstīt saprotamu komentāru\n"
            "5. Izrakstīt klientam aktu-rēķinu un norakstīt iekārtas",
            reply_markup=MAIN_MENU,
        )
        return

    if text == "Pults operatora darbības":
        await update.message.reply_text(
            "PULTS OPERATORA DARBĪBAS\n\n"
            "1. Pieņemt un operatīvi apstrādāt trauksmes signālu\n"
            "2. Novērtēt signāla pareizību un reaģēšanas plānu\n"
            "3. Sazināties ar klientu,par nepieciešamību pareizi reaģēt\n"
            "4. Nosūtīt Mobilo grupu",
            reply_markup=MAIN_MENU,
        )
        return

    if text == "Kontakti":
        await update.message.reply_text(
            "Izvēlieties kontaktu:",
            reply_markup=CONTACTS_MENU,
        )
        return

    if text in CONTACTS:
        await update.message.reply_text(
            CONTACTS[text],
            reply_markup=CONTACTS_MENU,
        )
        return

    if text == "Veidlapas":
        if not form_files:
            await update.message.reply_text(
                f"Nav veidlapu:\n{FORMS_DIR}",
                reply_markup=MAIN_MENU,
            )
            return

        await update.message.reply_text(
            "Izvēlieties veidlapu:",
            reply_markup=build_menu(form_files),
        )
        return

    if text in form_map:
        with open(form_map[text], "rb") as f:
            await update.message.reply_document(
                document=f,
                caption=f"Veidlapa: {text}",
                reply_markup=BACK_MENU,
            )
        return

    # Autoparka piekļuves pieprasījums
    if text == "Autoparks":
        if context.chat_data["autopark_access"]:
            if autopark_error:
                await update.message.reply_text(
                    autopark_error,
                    reply_markup=MAIN_MENU,
                )
                return

            await update.message.reply_text(
                "Izvēlieties tehniķi:",
                reply_markup=build_autopark_menu(button_map),
            )
            return

        context.chat_data["waiting_for_autopark_pin"] = True
        await update.message.reply_text(
            "Ievadiet Autoparka PIN kodu:",
            reply_markup=MAIN_MENU,
        )
        return

    # PIN pārbaude
    if context.chat_data["waiting_for_autopark_pin"]:
        if text == AUTOPARK_PIN:
            context.chat_data["waiting_for_autopark_pin"] = False
            context.chat_data["autopark_access"] = True

            if autopark_error:
                await update.message.reply_text(
                    autopark_error,
                    reply_markup=MAIN_MENU,
                )
                return

            await update.message.reply_text(
                "PIN pareizs. Izvēlieties tehniķi:",
                reply_markup=build_autopark_menu(button_map),
            )
            return
        else:
            context.chat_data["waiting_for_autopark_pin"] = False
            context.chat_data["autopark_access"] = False
            await update.message.reply_text(
                "Nepareizs PIN kods.",
                reply_markup=MAIN_MENU,
            )
            return

    if button_map and text in button_map:
        if not context.chat_data["autopark_access"]:
            await update.message.reply_text(
                "Nav piekļuves Autoparkam. Vispirms ievadiet PIN kodu.",
                reply_markup=MAIN_MENU,
            )
            return

        real_name = button_map[text]
        info_text = tech_info_map.get(real_name, "❌ Informācija nav atrasta.")

        await update.message.reply_text(
            f"🚗 {real_name}\n\n{info_text}",
            reply_markup=build_autopark_menu(button_map),
        )
        return

    if text == "⬅️ Atpakaļ uz galveno izvēlni":
        context.chat_data["waiting_for_autopark_pin"] = False
        await update.message.reply_text(
            "Galvenā izvēlne:",
            reply_markup=MAIN_MENU,
        )
        return

    await update.message.reply_text(
        "Izvēlieties sadaļu no izvēlnes.",
        reply_markup=MAIN_MENU,
    )


def main():
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))

    print("Bots palaists...")
    app.run_polling()


if __name__ == "__main__":
    main()